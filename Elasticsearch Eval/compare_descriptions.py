"""
Compare effectiveness of two dataset description variants using Elasticsearch.


"""

import csv
import json
import math
import warnings
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from openai import OpenAI
from elasticsearch import Elasticsearch
from elasticsearch.helpers import bulk
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

# ---------- CONFIG ----------
OPENAI_API_KEY = ""

#Run elasticsearch-9.3.3/bin/elasticsearch.bat in console for password
ELASTIC_PASSWORD = ""

ES_HOST = "https://localhost:9200"
ES_USER = "elastic"

ROOT = Path(__file__).parent
XLSX = ROOT / "Testing Data.xlsx"
QUERY_CACHE = ROOT / "queries_cache.json"

# Per-sheet data_dict CSVs to merge in by dataset_id (Education originally lacked
# the column in the xlsx; this CSV fills it in).
EXTRA_DATA_DICT_CSVS = {
    "Education": ROOT / "data_dict_results_education.csv",
}

QUERIES_PER_DATASET = 3
TOP_K = 10
LLM_MODEL = "gpt-4o-mini"
LLM_WORKERS = 4
LLM_MAX_RETRIES = 8

VARIANTS = {
    "new": "search_focused_description",
    "original": "search_focused_description_original",
}
METRIC_NAMES = ["Recall@1", "Recall@5", "Recall@10", "MRR", "NDCG@10"]
# ----------------------------


def load_data(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))
        records = []
        for r in rows:
            if r[0] is None:
                continue
            records.append(dict(zip(headers, r)))
        out[sheet] = records
    for sheet, csv_path in EXTRA_DATA_DICT_CSVS.items():
        if sheet in out and csv_path.exists():
            _merge_data_dict_csv(out[sheet], csv_path)
    return out


def _merge_data_dict_csv(records, csv_path):
    """Add `data_dict` field to records by joining on dataset_id."""
    by_id = {}
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            ds_id = row.get("dataset_id")
            if ds_id and ds_id not in by_id:
                by_id[ds_id] = row.get("data_dict") or ""
    matched = 0
    for d in records:
        dd = by_id.get(d["dataset_id"])
        if dd:
            d["data_dict"] = dd
            matched += 1
    print(f"  merged {matched}/{len(records)} data_dicts from {csv_path.name}")


def generate_queries_for_dataset(client, dataset):
    desc_a = (dataset.get("description") or "")[:1200]
    desc_b = (dataset.get("description_original") or "")[:1200]
    data_dict = (dataset.get("data_dict") or "")[:3000]

    prompt = (
        f"You are simulating a data analyst searching a NYC OpenData catalog to find a "
        f"specific dataset for analytical work. Generate exactly {QUERIES_PER_DATASET} "
        f"specific, technical search queries that reference field names, measures, "
        f"dimensions, time periods, geographic units, or aggregations the dataset clearly "
        f"supports. Avoid generic catalog phrases like 'dataset overview' or 'comprehensive "
        f"information.' Write queries the way an analyst types them.\n\n"
        f"Vary the three queries:\n"
        f"- One query naming a specific field or measurement (e.g. 'avg SAT math score by school').\n"
        f"- One query combining a field with a filter or breakdown (e.g. 'restaurant violations by borough 2019').\n"
        f"- One query phrased as an analytical question (e.g. 'leading causes of death across age groups').\n\n"
        f'Output JSON: {{"queries": ["...", "...", "..."]}}\n\n'
        f"Data dictionary (authoritative field names, may be empty):\n{data_dict}\n\n"
        f"Description A:\n{desc_a}\n\nDescription B:\n{desc_b}"
    )
    resp = client.chat.completions.create(
        model=LLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
        response_format={"type": "json_object"},
        temperature=0.7,
    )
    return json.loads(resp.choices[0].message.content)["queries"]


def generate_all_queries(all_data):
    cache = json.loads(QUERY_CACHE.read_text()) if QUERY_CACHE.exists() else {}
    todo = [
        d
        for datasets in all_data.values()
        for d in datasets
        if d["dataset_id"] not in cache
    ]
    total = sum(len(v) for v in all_data.values())
    if not todo:
        print(f"  All {total} datasets cached.")
        return cache

    print(f"  Generating queries for {len(todo)} datasets ({total - len(todo)} cached)...")
    client = OpenAI(api_key=OPENAI_API_KEY, max_retries=LLM_MAX_RETRIES)
    done = 0
    failures = 0
    with ThreadPoolExecutor(max_workers=LLM_WORKERS) as ex:
        futures = {ex.submit(generate_queries_for_dataset, client, d): d for d in todo}
        for f in as_completed(futures):
            d = futures[f]
            try:
                cache[d["dataset_id"]] = f.result()
            except Exception as e:
                failures += 1
                print(f"    failed {d['dataset_id']}: {e}")
                # Don't cache failures — leave the dataset out so a re-run retries it.
            done += 1
            if done % 25 == 0:
                print(f"    {done}/{len(todo)} ({failures} failures so far)")
                QUERY_CACHE.write_text(json.dumps(cache, indent=2))
    QUERY_CACHE.write_text(json.dumps(cache, indent=2))
    if failures:
        print(f"  {failures} datasets still missing queries; re-run the script to retry.")
    return cache


INDEX_BODY = {
    "mappings": {
        "properties": {
            "dataset_id": {"type": "keyword"},
            "description": {"type": "text", "analyzer": "english"},
        }
    }
}


def setup_index(es, name):
    if es.indices.exists(index=name):
        es.indices.delete(index=name)
    es.indices.create(index=name, body=INDEX_BODY)


def bulk_index(es, name, datasets, field):
    actions = []
    for d in datasets:
        text = d.get(field) or ""
        if not text:
            continue
        actions.append({
            "_index": name,
            "_id": d["dataset_id"],
            "_source": {
                "dataset_id": d["dataset_id"],
                "description": text,
            },
        })
    bulk(es, actions, refresh=True)


def search_bm25(es, index, query, size=TOP_K):
    res = es.search(
        index=index,
        size=size,
        query={"match": {"description": query}},
        _source=["dataset_id"],
    )
    return [hit["_source"]["dataset_id"] for hit in res["hits"]["hits"]]


def _empty_acc():
    return {"n": 0, "r1": 0, "r5": 0, "r10": 0, "mrr": 0.0, "ndcg": 0.0}


def _accumulate(acc, results, gold_id):
    acc["n"] += 1
    try:
        rank = results.index(gold_id) + 1
    except ValueError:
        return
    if rank <= 1:
        acc["r1"] += 1
    if rank <= 5:
        acc["r5"] += 1
    if rank <= 10:
        acc["r10"] += 1
    acc["mrr"] += 1.0 / rank
    acc["ndcg"] += 1.0 / math.log2(rank + 1)


def _finalize(acc):
    n = acc["n"]
    if n == 0:
        return {"queries": 0, **{m: 0.0 for m in METRIC_NAMES}}
    return {
        "queries": n,
        "Recall@1": acc["r1"] / n,
        "Recall@5": acc["r5"] / n,
        "Recall@10": acc["r10"] / n,
        "MRR": acc["mrr"] / n,
        "NDCG@10": acc["ndcg"] / n,
    }


def evaluate(es, index, datasets, queries):
    acc = _empty_acc()
    for d in datasets:
        ds_id = d["dataset_id"]
        for q in queries.get(ds_id, []):
            if not q:
                continue
            _accumulate(acc, search_bm25(es, index, q), ds_id)
    return _finalize(acc)


def print_table(label, n_queries, new_metrics, orig_metrics):
    print(f"\n{label} ({n_queries} queries)")
    print(f"  {'Metric':<12} {'new':>10} {'original':>10} {'delta':>10}")
    for m in METRIC_NAMES:
        nv, ov = new_metrics[m], orig_metrics[m]
        print(f"  {m:<12} {nv:>10.4f} {ov:>10.4f} {nv - ov:>+10.4f}")


def main():
    print("Loading data...")
    all_data = load_data(XLSX)
    for s, ds in all_data.items():
        print(f"  {s}: {len(ds)} datasets")

    print("\nGenerating queries...")
    queries = generate_all_queries(all_data)

    print("\nConnecting to Elasticsearch...")
    es = Elasticsearch(
        ES_HOST,
        basic_auth=(ES_USER, ELASTIC_PASSWORD),
        verify_certs=False,
        ssl_show_warn=False,
    )
    if not es.ping():
        raise SystemExit("Cannot connect to Elasticsearch. Is it running on " + ES_HOST + "?")

    results = {}
    for sheet, datasets in all_data.items():
        results[sheet] = {}
        for variant, field in VARIANTS.items():
            index_name = f"datasets_{sheet.lower()}_{variant}"
            print(f"\n[{sheet}/{variant}] indexing into {index_name}...")
            setup_index(es, index_name)
            bulk_index(es, index_name, datasets, field)
            print(f"[{sheet}/{variant}] evaluating...")
            results[sheet][variant] = evaluate(es, index_name, datasets, queries)

    print("\n" + "=" * 60)
    print("RESULTS")
    print("=" * 60)
    for sheet, variants in results.items():
        print_table(sheet, variants["new"]["queries"], variants["new"], variants["original"])

    overall = {v: {m: 0.0 for m in METRIC_NAMES} for v in VARIANTS}
    overall_n = 0
    for variants in results.values():
        n = variants["new"]["queries"]
        overall_n += n
        for v in VARIANTS:
            for m in METRIC_NAMES:
                overall[v][m] += variants[v][m] * n
    if overall_n:
        for v in VARIANTS:
            for m in METRIC_NAMES:
                overall[v][m] /= overall_n
    print_table("OVERALL (weighted)", overall_n, overall["new"], overall["original"])


if __name__ == "__main__":
    main()
